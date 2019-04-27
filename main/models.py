from django.db import models
from openpyxl import load_workbook

COHORT = (
    ('J', 'Jan Extended'),
    ('F', 'Feb Standard'),
    ('A', 'Aug Standard'),
    ('M', 'May Express'),
    ('O', 'October Express'),
)

YEAR = (
    ('18', '2018'),
    ('19', '2019'),
    ('20', '2020'),
    ('21', '2021'),
)

MODULE = (
    ('WC', 'Web Coding'),
    ('CE', 'Computer Essentials'),
    ('IS', 'Information Systems'),
    ('PR', 'Programming'),
    ('SI', 'Social Issues')
)


def populate_students(student_list_fn, cohort_choice, year_choice):
    wb = load_workbook(filename=student_list_fn)

    for ws in wb.sheetnames:

        if ws[2] == cohort_choice.lower():
            class_number = ws[-1]
            ws = wb.get_sheet_by_name(ws)
            # Get the names of the students
            for row in ws.iter_rows(min_row=3, max_col=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        raw_student_name = cell.value

                        lastname, givennames = raw_student_name.split(',')
                        try:
                            firstname, preferred = givennames.split('(')
                        except ValueError:
                            firstname, preferred = givennames, ""
                        Student.objects.get_or_create(
                            year_choices=year_choice,
                            cohort_choices=cohort_choice,
                            class_number=class_number,
                            firstname=firstname.strip(),
                            lastname=lastname.strip(),
                            preferred=preferred.replace(')', "").strip())


class Student(models.Model):
    year_choices = models.CharField(max_length=2, choices=YEAR)
    cohort_choices = models.CharField(max_length=1, choices=COHORT)
    class_number = models.CharField(max_length=1, choices=((str(i), str(i)) for i in range(1, 10)))
    firstname = models.CharField(max_length=130)
    lastname = models.CharField(max_length=130)
    preferred = models.CharField(max_length=130, blank=True, null=True)

    def __str__(self):
        return f'{self.lastname + ", "} {self.firstname} {"(" + self.preferred + ")" if self.preferred is not None else ""} '


class StudentList(models.Model):
    year_choices = models.CharField(max_length=2, choices=YEAR)
    cohort_choices = models.CharField(max_length=1, choices=COHORT)

    sl_xls = models.FileField(upload_to='student_list/%Y-%m-%d/%H/%M', verbose_name="Student List")
    uploaded_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        get_latest_by = 'uploaded_at'

    def _update_class_list(self, created=False):
        if created:
            populate_students(
                self.sl_xls.path,
                self.cohort_choices,
                self.year_choices,
            )

    def save(self, *args, **kwargs):
        created = self.pk is None
        super(StudentList, self).save(*args, **kwargs)
        self._update_class_list(created)

    def __str__(self):
        return f'{self.get_cohort_choices_display()} student list for' \
            f' {self.get_cohort_choices_display()} cohort, uploaded on {self.uploaded_at.strftime("%x")}'


class Scaffold(models.Model):
    year_choices = models.CharField(max_length=2, choices=YEAR)
    cohort_choices = models.CharField(max_length=1, choices=COHORT)
    module = models.CharField(max_length=2, choices=MODULE)
    pdf = models.FileField(upload_to='scaffold/%Y-%m-%d/%H/%M', verbose_name="Scaffold")
    uploaded_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        get_latest_by = 'uploaded_at'

    def __str__(self):
        return f'{self.get_cohort_choices_display()} scaffold for' \
            f' {self.get_cohort_choices_display()} cohort, uploaded on {self.uploaded_at.strftime("%x")}'
